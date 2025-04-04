import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import io
import numpy as np
import platform
import matplotlib.font_manager as fm
import os
import time
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP
import tempfile
import platform

import plotly.express as px
import plotly.graph_objects as go


__version__ = '0.2'
__updatedate__ = '2025.04.04'


# 设置中文字体
#plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'sans-serif']  # 用来正常显示中文
#plt.rcParams['axes.unicode_minus'] = False  # 正常显示负号


def show_about_page():
    """显示关于页面的信息"""
    st.title("关于")
    st.write("这是一个数据分析工具，以下是系统信息：")
    
    # 获取操作系统名称
    os_name = platform.system()
    st.write(f"**操作系统名称 (platform.system()):** {os_name}")

    # 获取操作系统版本
    os_release = platform.release()
    st.write(f"**操作系统发行版本 (platform.release()):** {os_release}")

    # 获取更详细的操作系统版本信息
    os_version = platform.version()
    st.write(f"**操作系统详细版本 (platform.version()):** {os_version}")

    # 获取平台架构 (例如 'x86_64', 'i386')
    machine_arch = platform.machine()
    st.write(f"**机器架构 (platform.machine()):** {machine_arch}")

    # 获取更全面的平台信息字符串
    platform_info = platform.platform()
    st.write(f"**完整平台信息 (platform.platform()):**")
    st.code(platform_info, language='text')

    # 获取 uname 信息 (类似 Linux 命令 `uname -a`)
    uname_info = platform.uname()
    st.write(f"**uname 信息 (platform.uname()):**")
    st.code(str(uname_info), language='text')


# 跨平台字体设置函数
def setup_chinese_font():
    # 不同平台的常用中文字体
    if platform.system() == 'Windows':
        font_list = ['SimHei', 'Microsoft YaHei', 'SimSun']
    elif platform.system() == 'Darwin':  # macOS
        font_list = ['PingFang SC', 'Heiti SC', 'STHeiti']
    else:  # Linux
        font_list = ['WenQuanYi Micro Hei', 'Noto Sans CJK SC', 'DejaVu Sans']
    
    # 获取系统可用字体
    available_fonts = [f.name for f in fm.fontManager.ttflist]
    
    # 查找第一个可用的中文字体
    for font in font_list:
        if font in available_fonts:
            plt.rcParams['font.sans-serif'] = [font, 'DejaVu Sans', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False
            return font
    
    # 如果没有找到中文字体，使用默认设置
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    return 'DejaVu Sans'

# 在模块加载时设置字体
setup_chinese_font()



# 设置页面配置
st.set_page_config(
    page_title="M2数据分析工具",
    page_icon="��",
    layout="wide"
)

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
def check_file_size(file):
    if file.size > MAX_FILE_SIZE:
        st.error(f"文件大小超过限制（最大50MB）")
        return False
    return True

def process_m2_data(file_content):
    """处理M2数据并生成图表"""
    # 找到Frame (Quantitative)部分
    # frame_part = file_content.split('Frame (Quantitative)')[1].strip()
    # 查找包含"Frame"的部分
    frame_markers = ['Frame (Quantitative)', 'Frame Results']
    frame_part = None

    for marker in frame_markers:
        if marker in file_content:
            frame_part = file_content.split(marker)[1].strip()
            break

    if frame_part is None:
        raise ValueError("未找到Frame数据部分，请检查文件格式")

    # 使用pandas读取这部分数据
    df = pd.read_csv(io.StringIO(frame_part))
    
    # 计算比值数据
    ratio = []
    for i in range(len(df)):
        x_val = df.iloc[i, 0]  # Beam Width X
        y_val = df.iloc[i, 1]  # Beam Width Y
        if y_val / x_val > 1:
            ratio.append(x_val / y_val)
        else:
            ratio.append(y_val / x_val)
    
    # 创建图形
    fig, ax1 = plt.subplots(figsize=(10, 6))
    
    # 绘制曲线
    ax1.plot(df.iloc[:, 2], df.iloc[:, 0], 'r-', label='Beam Width X (μm)')
    ax1.plot(df.iloc[:, 2], df.iloc[:, 1], 'b-', label='Beam Width Y (μm)')
    
    # 设置第一个Y轴的标签
    ax1.set_title('M2 foucs analysis', fontsize=14)
    ax1.set_xlabel('Z Location (mm)', fontsize=12)
    ax1.set_ylabel('Beam Width (μm)', fontsize=12)
    ax1.legend(loc='upper left', fontsize=10)
    ax1.grid(True)
    
    # 创建第二个Y轴用于比值
    ax2 = ax1.twinx()
    ax2.plot(df.iloc[:, 2], ratio, 'g--', label='beam roundness')
    
    # 添加合格标准线
    standard_line = ax2.axhline(y=0.9, color='orange', linestyle='-', linewidth=2, alpha=0.7, label='Qualification criteria(0.9)')
    
    # 添加合格区域填充
    x_range = df.iloc[:, 2]
    ax2.fill_between(x_range, 0.9, 1.0, color='lightgreen', alpha=0.3, label='qualified area')
    ax2.fill_between(x_range, 0.85, 0.9, color='pink', alpha=0.3, label='unqualified area')
    
    ax2.set_ylabel('beam roundness(after focus)', color='g', fontsize=12)
    ax2.tick_params(axis='y', colors='g')
    ax2.set_ylim(0.85, 1.0)
    ax2.legend(loc='upper right', fontsize=10)
    
    plt.tight_layout()
    return fig

def save_fig_to_bytes(fig):
    """将matplotlib图形保存为字节流"""
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
    buf.seek(0)
    return buf

def process_summary_data(template_file, data_files):
    """处理纠正预防措施汇总数据"""
    try:
        # 读取模板文件
        wb = load_workbook(template_file)
        ws = wb.active
        
        # 获取第四行的数据作为映射表
        first_row_data = {}
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell_value = ws.cell(row=4, column=col).value
            if cell_value is not None:
                first_row_data[col_letter] = cell_value
        
        # 从第5行开始处理数据文件
        row = ws.max_row + 1
        processed_files_count = 0
        
        for data_file in data_files:
            load_wb = load_workbook(data_file, data_only=True)
            load_sheet = load_wb.active
            
            # 写入序号和文件名
            ws[f'A{row}'] = row-5
            ws[f'B{row}'] = os.path.basename(data_file.name)
            
            # 根据映射表处理数据
            for target_cell, source_cell in first_row_data.items():
                if source_cell is not None:
                    source_value = load_sheet[source_cell].value
                    if isinstance(source_value, datetime):
                        source_value = source_value.strftime('%Y/%m/%d')
                    ws[f'{target_cell}{row}'] = source_value
            
            processed_files_count += 1
            row += 1
            load_wb.close()
        
        # 保存处理后的文件
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer, processed_files_count
        
    except Exception as e:
        st.error(f"处理文件时出错：{str(e)}")
        return None, 0

def round_to_decimal(value: str, decimal_places: int) -> float:
    """
    将字符串数值四舍五入到指定小数位数
    Args:
        value (str): 需要四舍五入的字符串数值
        decimal_places (int): 四舍五入后保留的小数位数
    Returns:
        float: 四舍五入后的浮点数
    """
    format_str = '0.' + '0' * decimal_places
    return float(Decimal(value).quantize(Decimal(format_str), rounding=ROUND_HALF_UP))

def process_product_data(template_file, data_files):
    """处理产成品数据汇总"""
    try:
        # 读取模板文件
        wb = load_workbook(template_file)
        ws = wb.active
        
        # 获取第一行的数据作为映射表
        first_row_data = {}
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell_value = ws.cell(row=1, column=col).value
            if cell_value is not None:
                first_row_data[col_letter] = cell_value
        
        # 从最后一行开始处理数据文件
        row = ws.max_row + 1
        processed_files_count = 0
        
        for data_file in data_files:
            load_wb = load_workbook(data_file, data_only=True)
            load_sheet = load_wb.active
            
            # 写入序号和文件名
            ws[f'A{row}'] = row-1
            ws[f'B{row}'] = os.path.basename(data_file.name)
            
            # 根据映射表处理数据
            for target_cell, source_cell in first_row_data.items():
                if source_cell is not None:
                    source_value = load_sheet[source_cell].value
                    cell = ws[f'{target_cell}{row}']
                    cell.value = source_value
                    
                    # 如果是日期类型，设置日期格式
                    if isinstance(source_value, datetime):
                        cell.number_format = 'yyyy/mm/dd'
                    
                    # 特殊处理AR列（功率范围）
                    if target_cell == 'AR':
                        var1, var2 = source_value.split('-')
                        var1 = int(var1[:-3])
                        var2 = int(var2[:-5])
                        ws[f'AR{row}'] = var1
                        ws[f'AS{row}'] = var2
                    
                    # 特殊处理AT列（功率计算）
                    elif target_cell == 'AT':
                        power = ws[f'AT{row}'].value
                        ws[f'AU{row}'] = round_to_decimal(power / var1 * 1000, 0)
            
            processed_files_count += 1
            row += 1
            load_wb.close()
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 只为新添加的行添加边框
        for row in range(ws.max_row - processed_files_count + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # 保存处理后的文件
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer, processed_files_count
        
    except Exception as e:
        st.error(f"处理文件时出错：{str(e)}")
        return None, 0



###### 日志相关处理函数 #################
### 报警日志 ####
def process_alarm_log(df):
    """处理报警日志数据"""
    try:
        # 确保时间列是datetime类型
        df['time'] = pd.to_datetime(df['time'], errors='coerce', format='mixed')
        df = df.dropna(subset=['time'])
        
        # 让用户选择报警类型列
        alarm_column = st.selectbox("选择报警类型列", df.columns)
        
        # 时间范围选择
        min_time = df['time'].min()
        max_time = df['time'].max()
       
        start_time = st.date_input("开始时间", min_time, key="alarm_start_time")
        end_time = st.date_input("结束时间", max_time, key="alarm_end_time")

        # 过滤数据
        mask = (df['time'].dt.date >= start_time) & (df['time'].dt.date <= end_time)
        filtered_df = df[mask]
        
        # 显示数据
        st.dataframe(filtered_df)
        
        # 报警统计
        st.subheader("报警统计")
        alarm_counts = filtered_df[alarm_column].value_counts()
        fig = px.bar(alarm_counts, title="报警类型分布")
        fig.update_layout(
            xaxis=dict(
                tickmode='auto',
                nticks=20,
                tickangle=45,
                tickformat='%Y-%m-%d %H:%M:%S'  # 时间格式
            )
        )
        st.plotly_chart(fig)

    except Exception as e:
        st.error(f"处理报警日志时出错：{str(e)}")
        st.write("请检查文件格式是否正确")


def process_operate_log(df):
    """处理操作日志数据"""
    try:
        # 确保时间列是datetime类型
        df['time'] = pd.to_datetime(df['time'], errors='coerce', format='mixed')
        df = df.dropna(subset=['time'])
        
        # 让用户选择操作类型列
        operate_column = st.selectbox("选择操作类型列", df.columns)
        
        # 时间范围选择
        min_time = df['time'].min()
        max_time = df['time'].max()

        start_time = st.date_input("开始时间", min_time, key="operate_start_time")
        end_time = st.date_input("结束时间", max_time, key="operate_end_time")

        # 过滤数据
        mask = (df['time'].dt.date >= start_time) & (df['time'].dt.date <= end_time)
        filtered_df = df[mask]
        
        # 显示数据
        st.dataframe(filtered_df)
        
        # 操作类型统计
        st.subheader("操作类型统计")
        operate_counts = filtered_df[operate_column].value_counts()
        fig = px.bar(operate_counts, title="操作类型分布")
        fig.update_layout(
            xaxis=dict(
                tickmode='auto',
                nticks=20,
                tickangle=45,
                tickformat='%Y-%m-%d %H:%M:%S'  # 时间格式
            )
        )
        st.plotly_chart(fig)

    except Exception as e:
        st.error(f"处理操作日志时出错：{str(e)}")
        st.write("请检查文件格式是否正确")



def process_status_log(df):
    """处理状态日志数据"""
    try:
        # 确保时间列是datetime类型
        df['time'] = pd.to_datetime(df['time'], errors='coerce', format='mixed')
        df = df.dropna(subset=['time'])
        
        # 时间范围选择
        min_time = df['time'].min()
        max_time = df['time'].max()
        start_time = st.date_input("开始时间", min_time, key="status_start_time")
        end_time = st.date_input("结束时间", max_time, key="status_end_time")
        
        # 过滤数据
        mask = (df['time'].dt.date >= start_time) & (df['time'].dt.date <= end_time)
        filtered_df = df[mask]
        
        # 参数选择
        st.subheader("选择要显示的参数")
        all_columns = [col for col in filtered_df.columns if col != 'time']  # 排除时间列
        selected_columns = st.multiselect("选择参数", all_columns, default=all_columns[:5])
        
        if selected_columns:
            # 显示数据
            st.dataframe(filtered_df[['time'] + selected_columns])
            
            # 绘制选中的参数趋势图
            st.subheader("参数趋势图")
            fig = go.Figure()
            for col in selected_columns:
                # 将NaN值替换为0
                y_values = filtered_df[col].fillna(0)
                fig.add_trace(go.Scatter(
                    x=filtered_df['time'], 
                    y=y_values, 
                    name=col,
                    mode='markers',  # 只显示数据点
                    marker=dict(size=6)  # 设置数据点大小
                ))
            fig.update_layout(
                title="参数趋势图",
                xaxis_title="时间",
                yaxis_title="数值",
                hovermode="x unified",
                height=800,  # 增加图表高度
                margin=dict(l=50, r=50, t=50, b=50),  # 调整边距
                xaxis=dict(
                    tickmode='auto',  # 自动设置刻度
                    nticks=20,  # 增加刻度数量
                    tickangle=45,  # 倾斜角度
                    tickformat='%Y-%m-%d %H:%M:%S'  # 时间格式
                )
            )
            st.plotly_chart(fig, use_container_width=True)  # 使用容器宽度
    except Exception as e:
        st.error(f"处理状态日志时出错：{str(e)}")
        st.write("请检查文件格式是否正确")




def show_log_analysis():
    """显示日志分析页面"""
    st.title("设备日志分析")
    
    # 创建标签页
    tab1, tab2, tab3 = st.tabs(["报警日志", "操作日志", "状态日志"])
    
    with tab1:
        st.subheader("报警日志分析")
        alarm_file = st.file_uploader("上传报警日志文件", type=['xlsx'], key="alarm_log")
        if alarm_file:
            df_alarm = pd.read_excel(alarm_file)
            process_alarm_log(df_alarm)
    
    with tab2:
        st.subheader("操作日志分析")
        operate_file = st.file_uploader("上传操作日志文件", type=['xlsx'], key="operate_log")
        if operate_file:
            df_operate = pd.read_excel(operate_file)
            process_operate_log(df_operate)
    
    with tab3:
        st.subheader("状态日志分析")
        status_file = st.file_uploader("上传状态日志文件", type=['xlsx'], key="status_log")
        if status_file:
            df_status = pd.read_excel(status_file)
            process_status_log(df_status)




def main():
    # 侧边栏
    with st.sidebar:
        st.title("功能选择")
        
        # 设置侧边栏和按钮样式
        st.markdown("""
            <style>
            /* 调整侧边栏宽度 */
            section[data-testid="stSidebar"] {
                width: 200px;
                background: linear-gradient(180deg, #f8f9fa 0%, #ffffff 100%);
                padding: 1rem;
            }
            
            /* 按钮样式 */
            .stButton > button {
                font-size: 1.1rem;
                padding: 0.7rem 1rem;
                margin: 0.5rem 0;
                border-radius: 12px;
                background: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%);
                color: white;
                border: none;
                transition: all 0.3s ease;
                width: 75%;
                margin-left: 10%;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                position: relative;
                overflow: hidden;
            }
            
            .stButton > button::before {
                content: '';
                position: absolute;
                top: 0;
                left: -100%;
                width: 100%;
                height: 100%;
                background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
                transition: 0.5s;
            }
            
            .stButton > button:hover {
                transform: translateY(-2px);
                box-shadow: 0 4px 8px rgba(79, 70, 229, 0.3);
            }
            
            .stButton > button:hover::before {
                left: 100%;
            }
            
            /* 标题样式 */
            .sidebar-title {
                font-size: 1.5rem;
                margin-bottom: 1.5rem;
                color: #1f1f1f;
                text-align: center;
                font-weight: 600;
                background: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }
            
            /* 作者信息样式 */
            .author-info {
                margin-top: 2rem;
                padding-top: 1rem;
                border-top: 1px solid rgba(0,0,0,0.1);
                font-size: 0.9rem;
                color: #666;
                margin-left: 10%;
                width: 75%;
                text-align: center;
            }
            
            .author-info p {
                margin: 0.3rem 0;
                opacity: 0.8;
            }
            
            /* 选中按钮样式 */
            .stButton > button:focus {
                background: linear-gradient(135deg, #4f46e5 0%, #4338ca 100%);
                box-shadow: 0 0 0 2px rgba(79, 70, 229, 0.3);
            }
            </style>
        """, unsafe_allow_html=True)
        
        # 使用按钮进行功能切换
        if st.button("M2数据二次处理", use_container_width=True):
            st.session_state.selected_function = "M2数据二次处理"
        if st.button("纠正预防措施汇总", use_container_width=True):
            st.session_state.selected_function = "纠正预防措施汇总"
        if st.button("产成品数据汇总", use_container_width=True):
            st.session_state.selected_function = "产成品数据汇总"      
        if st.button("设备日志分析", use_container_width=True):
            st.session_state.selected_function = "设备日志分析"
        if st.button("关于", use_container_width=True):
            st.session_state.selected_function = "关于"

        # 初始化选择的功能
        if 'selected_function' not in st.session_state:
            st.session_state.selected_function = "M2数据二次处理"
            
        # 添加作者信息和版本信息
        st.markdown(f"""
            <div class="author-info">
                <p>版本: V{__version__}</p>
                <p>日期: {__updatedate__}</p>
                <p>作者: Dr.Shi</p>
                <p>Email: 8582864@qq.com</p>
                <p>© 2025 版权所有</p>
            </div>
        """, unsafe_allow_html=True)
    
    # 主页面
    if st.session_state.selected_function == "M2数据二次处理":
        st.title("M2数据二次处理")
        st.write("请上传CSV文件进行处理")
        
        # 文件上传
        uploaded_file = st.file_uploader("选择或拖拽CSV文件", type=['csv'])
        
        if uploaded_file is not None:
            try:
                # 检查文件大小
                if not check_file_size(uploaded_file):
                    return
                
                # 读取文件内容
                file_content = uploaded_file.getvalue().decode('utf-8')
                
                # 处理数据并显示图表
                fig = process_m2_data(file_content)
                st.pyplot(fig)
                
                # 添加下载按钮
                img_bytes = save_fig_to_bytes(fig)
                st.download_button(
                    label="下载图表",
                    data=img_bytes,
                    file_name="beam_analysis.png",
                    mime="image/png"
                )
                
            except Exception as e:
                st.error(f"处理文件时出错：{str(e)}")
                st.write("请确保上传的文件格式正确，包含'Frame (Quantitative)'部分")
    
    elif st.session_state.selected_function == "纠正预防措施汇总":
        st.title("纠正预防措施汇总")
        
        # 模板文件上传
        st.subheader("1. 上传模板文件")
        template_file = st.file_uploader("请上传模板文件（Excel格式）", type=['xlsx'], key="template")
        
        # 数据文件上传
        st.subheader("2. 上传需要汇总的文件")
        data_files = st.file_uploader("请上传需要汇总的文件（Excel格式）", type=['xlsx'], accept_multiple_files=True, key="data")
        
        if template_file and data_files:
            if st.button("开始处理"):
                with st.spinner("正在处理文件..."):
                    start_time = time.time()

                    with tempfile.TemporaryDirectory() as temp_dir:
                        # 处理文件
                        # 临时目录会在with块结束时自动删除
                        output_buffer, processed_count = process_summary_data(template_file, data_files)
                        
                        if output_buffer:
                            # 显示处理结果
                            st.subheader("3. 处理结果")
                            st.success(f"成功处理 {processed_count} 个文件")
                            
                            # 显示处理时间
                            elapsed_time = time.time() - start_time
                            st.info(f"处理用时: {elapsed_time:.2f} 秒")
                            
                            # 提供下载按钮
                            st.download_button(
                                label="下载汇总文件",
                                data=output_buffer,
                                file_name="纠正预防措施汇总表.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error("处理失败，请检查文件格式是否正确")                

    
    elif st.session_state.selected_function == "产成品数据汇总":
        st.title("产成品数据汇总")
        
        # 模板文件上传
        st.subheader("1. 上传模板文件")
        template_file = st.file_uploader("请上传模板文件（Excel格式）", type=['xlsx'], key="product_template")
        
        # 数据文件上传
        st.subheader("2. 上传需要汇总的文件")
        data_files = st.file_uploader("请上传需要汇总的文件（Excel格式）", type=['xlsx'], accept_multiple_files=True, key="product_data")
        
        if template_file and data_files:
            if st.button("开始处理"):
                with st.spinner("正在处理文件..."):
                    start_time = time.time()
                    
                    with tempfile.TemporaryDirectory() as temp_dir:
                        # 处理文件
                        # 临时目录会在with块结束时自动删除
                        output_buffer, processed_count = process_product_data(template_file, data_files)
                        
                        if output_buffer:
                            # 显示处理结果
                            st.subheader("3. 处理结果")
                            st.success(f"成功处理 {processed_count} 个文件")
                            
                            # 显示处理时间
                            elapsed_time = time.time() - start_time
                            st.info(f"处理用时: {elapsed_time:.2f} 秒")
                            
                            # 提供下载按钮
                            st.download_button(
                                label="下载汇总文件",
                                data=output_buffer,
                                file_name="产成品数据汇总表.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error("处理失败，请检查文件格式是否正确")

    # 日志分析页面
    elif st.session_state.selected_function == "设备日志分析":
        show_log_analysis()





    # 添加关于按钮
    elif st.session_state.selected_function == "关于":  
        show_about_page()







if __name__ == "__main__":
    main()